from openpyxl import load_workbook
class reed_xlsx():
    def __init__(self, dir_file_name: str, sheet_name = " "):
        """
        description : metodo que nos permite incializar el documento el cual vamos a leer

        params :
            - dir_file_name (str) : direccion y nombre del archivo donde esta alojado
            - sheet_name (str) : establece el nombre de la hoja que quieres leer
        """
        self.dir_file_name = dir_file_name
        self.wb = load_workbook(dir_file_name)
        if sheet_name in self.wb:
            self.ws = self.wb[sheet_name]


    def get_keys(self):
        """
        description : metodo que nos permite obtener los nombres de las columnas de la hoja

        params :
            - None

        return : keys (list) : lista con los nombres de las columnas de la hoja
        """
        keys = []
        for i in range(self.ws.max_column):
            keys.append(self.ws.cell(row=2, column=i+1).value)
        return keys


    def get_values(self):
        """
        description : metodo que nos permite obtener los valores de las filas de la hoja

        params :
            - None

        return : values (list) : lista con los valores de las filas de la hoja
        """
        # lista que contiene las listas de datos de cada uno de los acreditantes 
        values = []
        # iteramos por el numero de acreditantes 
        for person in range(len(self.get_acreditantes())):
            # lista que contiene los datos de los acreditantes
            data_person = []
            #integramos las listas lde datos de cada uno de los acreditantes en una sola lista
            values.append(data_person)
            # iteramos por el numero de columnas para obtener los datos de las celdas  
            for i in range(self.ws.max_column):
                #empezamos por obtener los datos 3 filas y 2 columnas a lo incial 
                data_person.append(self.ws.cell(row=person+3, column=i+2).value)
        
        return values


    def get_acreditantes(self):
        """
        description : metodo que nos permite obtener los nombres de los acreditantes de la hoja

        params :
            - None

        return : list
        """
        # lista de los nombres de los acreditantes 
        acreditantes  = []
        # iteramos por el rango de las columnas para obtener los acreditantes 
        for i in range(self.ws.max_row):
            # agregamos a la lista cada uno de los acreditantes 
            acreditantes.append(self.ws.cell(row=i+1, column=1).value)
        return acreditantes[2:]


    def get_data_row(self)  :
        # se obtiene todos los nombre de la columna menos el primero pues no lo necesitamos
        # la primera columna de datos contiene los nombres de los acreditantes
        
        keys = self.get_keys()[1:]
        values = self.get_values()
        acreditantes = self.get_acreditantes()

        def structured_data(data : list, acreditante: str):
            data_acreditantes  = {key : value for key, value in zip(keys, data)}
            return data_acreditantes

         
        return list(
            map(
                structured_data, 
                values,
                acreditantes
                )
            )

if __name__ == "__main__":
    work_file_name = "BASE DE DATOS CIRCULO DE CREDITO"
    sheet_name = "Info_creditos"
    creditos = reed_xlsx(f"./{work_file_name}.xlsx", sheet_name)
    data_acrediatnte = creditos.get_data_row()[0]
    for key,value in data_acrediatnte.items():
        print(f"{key}  : {value}")