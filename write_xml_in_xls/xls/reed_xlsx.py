from openpyxl import load_workbook
import pandas as pd
import os
import sys
from functools import reduce

sys.path.append("../")
from utils import maybe

class Read_xlsx():
    
    def __init__(self, file_path : str, worksheet_name : str):
        if self._is_valid_path(file_path):
            try : 
                self.file_path = file_path   
                self.wb = load_workbook(file_path)
            
                self.set_work_sheet(worksheet_name)
                self.df = self.__get_data_frame__(file_path)
            except KeyError :
                raise KeyError(f"la worksheet : {worksheet_name},  no existe")
        else:
            raise ValueError("Introduce una ruta o nombre de archivo valido")
    
    
    def  __str__(self) -> str:
        return f"""
    wb path : {self.file_path}
    ws set  : {self.ws}
        """
                
    
    def set_work_sheet(self, sheet_name : str):
        try :
            if sheet_name in self.wb.sheetnames:
                self.ws = self.wb[sheet_name]
        except ValueError:
            pass
    
    
    def _is_valid_path(self, path : str):
        """Descripcion : metodo que nos permite validar si un string es una ruta de una archivo

        Params:
            path (str): String que se quiere validar si es una ruta de una archivo

        Returns:
            boolean: retorna True o False dependiendo de la validacion 
        """
        if os.path.isabs(path) :
            return True
        elif not os.path.isabs(path) :
            return True
        else:
            return False
        
    
    def __get_data_frame__(self, path):
        return pd.read_excel(path)
    

    def get_column(self, column : str):
        return self.df[column]
    
    def filter_Tipe(self, column : str):
        """
        Descripcion : Metodo que nos permite obtener dada una condicion un nuevo df con tadas las filas que cumplieron esa condicion (esta funcion es un clousure).

        dicha condicion se basa en que el elemento de columna sea igual al de la condicion 

        params :
            - column (str) :Nombre de la columna que se tomara para evaual cada uno de sus datos en base a una condicion.
            - type (str | int) : Condicion la cual sera aplicada a cada uno de los datos en la columna 

        return : (df) df con las filas que cumplieron con dicha condicion
        """
        return lambda type : self.df[self.df[column] == type]
        
        
class Process_xlsx(Read_xlsx):
    
    def __init__(self,path, ws_name):
        super().__init__(path,ws_name)

    def get_tipos_operaciones(self):
        self.clv_prod_serv= "Clave de producto o servicio."
        tipo = "Tipo."
        return set(self.get_column(self.clv_prod_serv))
    

    def get_clasificacion_de_operacion(self):
        operaciones = self.get_tipos_operaciones()
        column = self.filter_Tipe(self.clv_prod_serv)
        df = reduce(
            lambda acc_df, current_df : pd.concat([acc_df, current_df], axis=0), list(map(column, operaciones))
            )
        print(df)
        self.sort_df = df
        df.to_excel(self.file_path,  sheet_name='Clasificacion de gastos', index=False)

    


def main():
    wb_path_e = "Enero_egresos 1.xlsx"
    wb_path_i = "Enero_ingresos_2.xlsx"
    ws_name = "Conjunto de gastos 2."
    wb = Process_xlsx(wb_path_e, ws_name)
    wb.get_clasificacion_de_operacion()



if __name__ == '__main__':
    main()