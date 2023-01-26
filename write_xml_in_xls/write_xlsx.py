from openpyxl import load_workbook
from openpyxl import Workbook
from typeguard import typechecked
from functools import reduce
from subprocess import run, PIPE
import os
import copy


class write_xlsx():
    """
    description: clase la cual nos proporciona metodo para la escritura y manipulacion de archivos excel

    attributes:
        path_file (str) : excel document path
        Excel_document (load_workbook): instancia de la clase load_workbook con el doc excel
        sheet_name ( str ) : nombre de la hoja para manipular



    """
    
    sheets = {}
    item = 1
    
    @typechecked
    def __init__(self, file_name : str) -> None:

        """
        description: crea un nuevo documento con el cual se va a trabajar o establece el document que se manipulara.

        si deseas cargar un archivo ya existente pasarias el argumento de file_path o si desear crear uno nuevo el argumento de name_woorkbook

        param:
            - name_woorkbook (str) : Nombre del documento que se va a crear o trabajar
            - file_path (str) : excel document path


        """
                    
        if self._is_valid_path(file_name):
            
            def create_document():
                self._create_file(file_name)
                self.wb.save(file_name)
                
            # self._load_file(file_name) if os.path.exists(file_name) else create_document()
            if os.path.exists(file_name):
                self._load_file(file_name)
            else:
                create_document()
                self._load_file(file_name)
                
                
            self.__path_file = file_name
            return
            
        else:
            raise ValueError("Introduce una ruta o nombre de archivo valido")
                
    
    def _is_valid_path(self, path : str):
        """Descripcion : metodo que nos permite validar si un string es una ruta de una archivo

        Params:
            path (str): String que se quiere validar si es una ruta de una archivo

        Returns:
            boolean: retorna True o False dependiendo de la validacion 
        """
        if os.path.isabs(path) :
            return True
        elif not os.path.isabs(path) :
            return True
        else:
            return False

        self.wb.save(self.__path_file)

    def _load_file(self, file_name  :str):
        """Description : Carga el documento excel a travez de la ruta introducida

        Args:
            file_name (str): ruta del archivo que se quiere cargar
        """
        self.wb = load_workbook(file_name)
            
            
    def _create_file(self, file_name :str):
        """Descripción : Crea un documento excel a través de la ruta introducida 
        
        Args:
            file_name (str) : Ruta y nombre donde se quiere crear el documento excel
        """
        self.wb =  Workbook(file_name)

    def get_shenames(self) -> list[str, str]:
        """
        description:
            retorna una lista con los nombres de las hojas que contiene el documento introducido

        return (list) : lista con los nombres de las hojas del documento
        """
        print(self.wb.sheetnames)
        return self.wb.sheetnames


    @typechecked
    def set_sheet_name(self, sheet_name: str,) -> str:
        """
        description: establece la hoja del documento que se procesara o manipulara

        params:
            - sheet_name (str) : nombre de la hoja en el documento

        return (str) : nombre de la hoja en el documento
        """
        self.sheet_name = self.wb[sheet_name]
        return self.sheet_name
    
    
    def create_new_sheet(self, name_sheet : str):
            self.wb.active
            self.wb.create_sheet(name_sheet)
            self.wb.save(self.__path_file)
    

    def write_row_by_range(self, name_sheet :str, initial_cell : str, data = list, ):
        
        """
        descripcion : metodo que nos permite modficar una fila a tarves de una lista con datos que queremos que sean asignados
        a las celdad de forma ordenda 

        al pasar la lista se toma la longitud de esta para saber cuantos datos a partir de la ceda dada como parametro "initial_cell" se necesitan modificar e iterar por la misma asignando a cada celda un dato de la lista

        params :
            - name_sheet (str) : establece la hoja o crear una nueva donde se ballan a escribir datos 
            - initial_cell (str) : celda de la cual quieres partir la insercion de datos (ejemplo : "A1","B3" etc. )
            - data (list) : datos los cuales quieres insertar en la fila
        """
        def letra_a_numero(letra):
            if len(letra) == 1:
               return ord(letra) - ord('A') + 1
            else:
                return 26 * letra_a_numero(letra[:-1]) + letra_a_numero(letra[-1])
            
        def asigacion_de_data(data,cell):
            cell.value = data
            return cell
           
        #obtenemos la columna desde que se empezada
        colunn = initial_cell[0]
        
        #obtenemos la fila a editar
        row = int(initial_cell[-1]) 

        #obtenemos el numero de la culumna donde esta esa celda
        initial_column = letra_a_numero(colunn)
        
        #obtenemos el maximo de celdad que se van a escribir
        max_col = initial_column + len(data) -1
        
        #si el valor del argumento nombre existe dentro del workbook 
        if name_sheet in self.wb.sheetnames:
            #asignando el objeto worksheet a una propiedad 
            self.ws = self.wb[name_sheet]
        else:
            print("Create")
            #si no creamos una hoja con el nombre del argumento 
            # self.wb.create_sheet(name_sheet)
            # self.ws = self.wb.active
            # print(self.ws.name)
            self.ws = self.wb.active
            self.ws.title = name_sheet
                        
        # celda inicial y celda final del rango a escribir
        def rows_iter(row, initial_column, max_col):
            fila_a_editar = self.ws.iter_rows(
                min_row=row, 
                max_row=row, 
                min_col=initial_column, 
                max_col=max_col
                )
            
            # retornamos la validacion de la fila vacia
            fila_validada = self.Empty_row(fila_a_editar)
            
            return fila_validada if fila_validada else rows_iter(row + 1, initial_column, max_col)

        fila_a_editar = rows_iter(row, initial_column, max_col)
        print(fila_a_editar)    
        fila_a_editar = fila_a_editar[0]
        print(fila_a_editar)    
                
        value = list(
            map(asigacion_de_data, 
            data,
            fila_a_editar
        ))
        self.wb.save(self.__path_file)
        
    
    def Empty_row (self, fila_a_editar):        
        copi_row = copy.copy(list(fila_a_editar))
        check_empty = lambda row: reduce(lambda x, y: x or y.value, row, False)
        return copi_row if not any(map(check_empty, copi_row)) else False
    
    
    def delete_wb(self):
        #TODO : Encontrar el porque falla esta madre 
        print(self.__path_file)
        run(["rm", "-r", f"{self.__path_file}"])
        

if __name__ == "__main__":
    enero = write_xlsx("./xls/Enero_2022.xlsx")
    table =  [[1,2],[1,2]]
    for element in table:
        enero.write_row_by_range("Clasificacion de gastos", "A2",element)
