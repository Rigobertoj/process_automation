from openpyxl import Workbook


wb = Workbook()
ws = wb.active

ws_1 = wb.create_sheet("Clasificacion de gastos", 0)
ws_1["A1"] = "Fecha estado de cuenta"

print(ws_1["A1"].value)

cell_range = ws["A1" : "A10"]
print(cell_range)

wb.save("Enero_2022.xlsx")

# class Humano():
#     __name = str
     
#     def __init__(self, name) -> None:
#         self.__name = name

# yo = Humano("Rigo")

# print(yo.__name)

for row in ws_1.iter_rows(min_col=0, max_col=20, min_row=2, max_row=2):
   for cell in row:
       print(cell)


def obtener_numero(letra):
    if len(letra) == 1:
        return ord(letra) - ord('A') + 1
    else:
        return 26 * obtener_numero(letra[:-1]) + obtener_numero(letra[-1])

print(obtener_numero("Ñ"))