import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os

class write_xlsx():
    """
    description: clase la cual nos proporciona metodo para la escritura y manipulacion de archivos excel

    attributes:
        path_file (str) : excel document path
        Excel_document (load_workbook): instancia de la clase load_workbook con el doc excel
        sheet_name ( str ) : nombre de la hoja para manipular



    """
    sheets = {}


    def __init__(self, name_woorkbook = "", path_file = "",  ) -> None:

        """
        description: crea un nuevo documento con el cual se va a trabajar establece el document que se manipulara.

        si deseas cargar un archivo ya existente pasarias el argumento de file_path o si desear crear uno nuevo el argumento de name_woorkbook

        param:
            - name_woorkbook (str) : Nombre del documento que se va a crear o trabajar
            - file_path (str) : excel document path


        """
        try :
            print("try")
            if name_woorkbook  != "":
                self.__wb = Workbook(name_woorkbook)
                self.__path_file = path_file
                return
            elif path_file != "" and os.path.exists(path_file):
                self.__wb = load_workbook(path_file)
                self.__path_file = path_file
                return
        except: ValueError("Introduce una ruta o nombre de archivo valido")




    def get_shenames(self) -> list[str, str]:
        """
        description:
            retorna una lista con los nombres de las hojas que contiene el documento introducido

        return (list) : lista con los nombres de las hojas del documento
        """
        print(self.__wb.sheetnames)
        return self.__wb.sheetnames



    def set_sheet_name(self, sheet_name: str,) -> str:
        """
        description: establece la hoja del documento que se procesara o manipulara

        params:
            - sheet_name (str) : nombre de la hoja en el documento

        return (str) : nombre de la hoja en el documento
        """
        self.sheet_name = self.__wb[sheet_name]
        return self.sheet_name


    def write_row_by_range(self,name_sheet :str, initial_cell : int, data = list, ):
        """
        descripcion : metodo que nos
        """
        def letra_a_numero(letra):
            if len(letra) == 1:
               return ord(letra) - ord('A') + 1
            else:
                print("else")
                return 26 * letra_a_numero(letra[:-1]) + letra_a_numero(letra[-1])

        #obtenemos la columna desde que se empezada
        colunn = initial_cell[0]

        #obtenemos el numero de la culumna donde esta esa celda
        num_column = letra_a_numero(colunn)

        #obtenemos la fila a editar
        row = int(initial_cell[-1])

        #obtenemos el maximo de celdad que se van a escribir
        max_col = num_column + len(data) -1

        if name_sheet in self.__wb.sheetnames:
            ws = self.__wb[name_sheet]
        else:
            ws = self.__wb.create_sheet(name_sheet)

        # celda inicial y celda final del rango a escribir
        fila_a_editar = ws.iter_rows(
            min_row=row, max_row=row, min_col=num_column, max_col=max_col
            )

        def asigacion_de_data(value, cell):
            cell.value = value
            return cell

        list(
            map(asigacion_de_data, 
            data,
            tuple(
                map(lambda row : row, fila_a_editar))[0]
            )
        )

        self.__wb.save("./Enero_2022.xlsx")


enero = write_xlsx(path_file="./Enero_2022.xlsx")
enero.write_row_by_range("Clasificacion de gastos", "A1", [""])
