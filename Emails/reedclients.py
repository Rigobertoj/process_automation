from openpyxl import load_workbook
Path = "clients/clients.xlsx"

class reedClient():
    """
    attributes:
        path_file (str) : excel document path
        Excel_document (load_workbook): instancia de la clase load_workbook con el doc excel
        sheet_name ( str ) : nombre de la hoja para manipular

    description:
        clase la cual nos proporciona metodo para la leectura y manipulacion de archivos excel

    """
    def __init__(self,file_path: str, ) -> None:
        """
        param: 
            file_path (str) : excel document path

        description:
            establece el document para su manipulacion 
        """

        self.path_file = file_path
        self.Excel_document = load_workbook(file_path)

    
    def get_shenames(self) -> list[str]:
        """
        description:
            retorna la lista con los nombres de las hojas en el documento

        return list : lista con los nombres de las hojas del documento
        """
        print(self.Excel_document.sheetnames)
        return self.Excel_document.sheetnames


    def set_sheet_name(self, sheet_name: str,) -> str:
        """
        params: 
            sheet_name (str) : nombre de la hoja en el documento

        description:
            establece la hoja con la que se va atrabajar en el documento

        return str : nombre de la hoja en el documento
        """
        self.sheet_name = self.Excel_document[sheet_name]
        return self.sheet_name

    def get_colum(self, name_colum: str,) -> list:
        """ 
        params: 
            name_colum (str) : nombre de la columna la cual se quiere obtener
        
        description:
            metodo que nos permite obtener cualquier columna de la hoja de excel a travez del nombre de la columna o su titulo, siempre y cuando dicha columna este cargada en el documento excel

        return list : lista con los objetos de cada celda en la columna 
        
        """
        for colum in list(self.sheet_name.columns):
            for cell in colum:
                if name_colum in str(cell.value):
                
                    column_celda = list(colum)
                    print(name_colum)
                    return column_celda


    def delete_cell(self, tittle: str, data_list: list) -> list:
        """
        params: 
            tittle (str): nombre de el titulo de la columa
            data_list (list): lista con los objetos de la culumn 

        description: 
            Metodo que nos permite eliminar una cellda de la columna. La lista debe de estar compueta por los objettos de tipo cell \n utilicece para eliminar el titulo de la columna

        return list : lista con los valores de los objetos sin el titulo o la celada a eliminar 
        """
        new_lits = [item.value for item in data_list if item.value != tittle]
        return new_lits

                    
    def get_values(self, data_list:list) -> list:
        """
        params:
            column (list): lista con los objetos cell

        description:
            obten los valores de la lista con los objetos de tipo cell

        return list : una lista con los valores de los datos de tipo cell
        """
        data_list = [cell.value  for cell in data_list]
        return data_list

    
    def save(self,):
        """
        description:
            Guarda los cambios en el documento. Este debe de estar sin abrir para que los cambios se guarden correctamente
        """
        self.Excel_document.save(self.path_file)

if __name__ == "__main__":
    # clients = reedClient(Path)
    # clients.get_shenames()
    # clients.set_sheet_name("clients credito")
    pass

