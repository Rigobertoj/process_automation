from openpyxl import load_workbook
Path = "./clients.xlsx"

class reedClient():
    
    def __init__(self,file_path: str, ):
        self.path_file = file_path
        self.Excel_document = load_workbook(file_path)

    
    def get_shenames(self) -> list[str]:
        """retorna una lista con el nombre de las hojas en el documento"""
        print(self.Excel_document.sheetnames)
        return self.Excel_document.sheetnames

    def set_sheet_name(self, sheet_name: str,):
        """establece la hoja con la que se va atrabajar en el documento"""
        self.sheet_name = self.Excel_document[sheet_name]
        return self.sheet_name

    
    def save(self,):
        """
        Guarda los cambios en el documento
        este de de estar sin abrir para que los cambios se guarden correctamente
        """
        self.Excel_document.save(self.path_file)

clients = reedClient(Path)
clients.get_shenames()
clients.set_sheet_name("clients credito")


