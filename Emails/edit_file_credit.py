from pydoc import cli
from typing import List
from xml.dom.minidom import Element
import openpyxl

PATH = "./Control de crédito mensual .xlsx"



# clase  que nos permite modificar las celdas de TIIE
class TIIE_file_edit_from_py ():
    """
    atributos: 
        path_file (str) :  ruta por la cual se accedera y modificar el archivo excel

    class :
        clase la cual nos permite modificar y editar las secciones donde se hubica la tiie en un doc excel

        se empiza estableciendo la hoja con la cual se va atrabajar (set_shename)

        
    """
    def __init__(self, path_file: str):
        """
        constructor
        params: 
            path_file (str) : ruta por la cual se accedera y modificar el archivo excel
        """
        #se define la ruta a la cual se accede 
        self.path_file = path_file
        # se instancia una clase de openpyxl para cargar el archivo de la ruta
        self.Excel_document = openpyxl.load_workbook(path_file)


    #retorna una lista de las hojas existentes en un documento en excel
    def get_sheet_names(self) -> list:
        """
        metod:
            nos permite visualizar toda la lista de hojas de trabajo que tengamos en nuestro archivo excel

        retorn list : lista de las hojas existentes en el documento en excel
        
        """
        print(self.Excel_document.sheetnames)
        return self.Excel_document.sheetnames


    #establece la hoja con la que se va atrabajar
    def set_sheet_name(self, sheet_name: str):
        """
        params : 
            sheet_name (str): nombre de la hoja con la cual se quiere trabajar

        metod: 
            establace la hoja con la cual se va atrabar y tratar los datos
        """

        self.sheet_names = self.Excel_document[sheet_name]


    #retorna el mismo elemento pero convertido en un estring
    def conver_string(self,objeto_a_convertir):
        """
        params:
            objeto_a_convertir (any) : valores que se pretenden convertir en strings

        metod:
            que nos permite convertir un elemento a un string
        """
        return f"{objeto_a_convertir}"



     #retorna del objeto CELL la posicion donde esta la celda TIIE en una cadena de texto para su manipulacion 

    def _get_cell_as_string(self,object):
        """
        
        """
        element = self.conver_string(object)
        list_with_cell = element.split(".")
        cell = list_with_cell[1][0:-1]
        return cell


    #retorna un diccionario con un valor y una lista
    #el valor es la celda que almacena la columna de la TIIE
    def _get_column_TIIE(self): 
        self.colum = {
            "value": 0,
            "list": list
        }
        for row in list(self.sheet_names.rows):
            for cell in row:
                if cell.value == "TIIE":
                    self.colum["list"] = row
                    self.colum["value"] = cell  
        return self.colum


    #retorna un diccionario con un indice
    #EL INDICE es la posicon de la columna donde esta la TIIE
    def _get_index_column_TIIE(self):
        self.data_colum = {
            "index": int,
        }

        colum = self._get_column_TIIE()
        value_list = colum["list"]
        for i in range(len(value_list)):
            if(colum["value"] == value_list[i]):
                self.data_colum["index"] = i

        return self.data_colum

    # retorna dos listas
    # LA PRIMERA LIST es una lista de celdas las cuales tiene un valor asignado de TIIE
    # LA SEGUNDA LIST es una lista con las celdas las cuales no tiene un valor asignado

    def get_cell_TIIE(self):
        index_colum_TIIE = self._get_index_column_TIIE()
        values = []
        Cells_empty = []
        for cell in list(self.sheet_names.columns)[index_colum_TIIE["index"]]:
            if cell.value != None:
                values.append(cell)
            else:
                Cells_empty.append(cell)
        
        lists = {
        "Cells_whit_TIIE":values,
        "Cells_empty":Cells_empty
        } 

        return lists

        #retorna la siguiente celda en la clumna de las TIIE la cual este vacia
        
    def get_next_cell_empty(self):
        lists_cell_TIIE = self.get_cell_TIIE()
        list_cell_with_values_TIIE = lists_cell_TIIE["Cells_whit_TIIE"]
        last_cell_with_TIIE = list_cell_with_values_TIIE[-1]
        cell = self._get_cell_as_string(last_cell_with_TIIE)
        [letter ,cord_x, cord_y] = cell
        cord_y = int(cord_y)+1
        cell = letter + cord_x + str(cord_y)
        return cell

        # inserta un valor float en la siguiente celda de la column TIIE 
        
    def insert_TIIE_in_client(self,value: float, path_to_save: str):
        cell = self.get_next_cell_empty()
        self.sheet_names[cell] = value
        self.Excel_document.save(self.path_file)
        print(self.sheet_names[cell].value)


client = TIIE_file_edit_from_py(PATH)
client.set_sheet_name("Alejandro Ochoa")

# client.conver_string(1)
cord = client.insert_TIIE_in_client(.9,"D:/desarrolloDeSofware/backend/python/TIIE/read_xlsx/Control de crédito mensual .xlsx")


