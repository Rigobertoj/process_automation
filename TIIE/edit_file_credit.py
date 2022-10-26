from pydoc import cli
from typing import List
from xml.dom.minidom import Element
import openpyxl
from convert import conver_value

PATH = "./Control de crédito mensual .xlsx"



# clase  que nos permite modificar las celdas de TIIE
class TIIE_file_edit_from_py ():
    """
    atributos: 
        path_file (str) :  ruta por la cual se accedera y modificar el archivo excel
        Excel_document : instancia que permite manipular y modificar el doc excel
    class :
        clase la cual nos permite modificar y editar las secciones donde se hubica la tiie en un doc excel

        se empiza estableciendo la hoja con la cual se va atrabajar (set_shename)

        
    """
    def __init__(self, path_file: str):
        """
        constructor
        params: 
            path_file (str) : ruta por la cual se accedera y modificar el archivo excel
        """
        #se define la ruta a la cual se accede 
        self.path_file = path_file
        # se instancia una clase de openpyxl para cargar el archivo de la ruta
        self.Excel_document = openpyxl.load_workbook(path_file)
        self.conver_value = conver_value()


    #retorna una lista de las hojas existentes en un documento en excel
    def get_sheet_names(self) -> list:
        """
        metod:
            nos permite visualizar toda la lista de hojas de trabajo que tengamos en nuestro archivo excel

        retorn list : lista de las hojas existentes en el documento en excel
        
        """
        print(self.Excel_document.sheetnames)
        return self.Excel_document.sheetnames


    #establece la hoja con la que se va atrabajar
    def set_sheet_name(self, sheet_name: str):
        """
        params : 
            sheet_name (str): nombre de la hoja con la cual se quiere trabajar

        metod: 
            establace la hoja con la cual se va atrabar y tratar los datos
        """

        self.sheet_names = self.Excel_document[sheet_name]


    #retorna del objeto CELL la posicion donde esta la celda TIIE en una cadena de texto para su manipulacion 
    def get_cell_as_string(self,cell_object):
        """
        params:
            cell (<Cell>) : es el tipo de dato cell que biene en openpyxl

        description:
            nos permite retornar del objeto CELL la cordenada o posicion del objeto eliminando el tipo, nombre de la hoja y los signos <> 
        """
        #?objeto cell -> <Cell 'sheet name'.H15>

        #convertimos en un estring el objeto CELL 
        element = self.conver_value.conver_string(cell_object)
        
        #separamos la cadena apartir del punto que viene en la cadena 
        list_with_cell = element.split(".")

        #retornamos el segundo elemento que es la celda y quitamos el signo >
        #que esta en la ultima posicion del string
        cell = list_with_cell[1][0:-1]
        return cell


    #retorna un diccionario con un valor y una lista
    #el valor es la celda que almacena la columna de la TIIE
    def get_row(self, name_column: str ):
        """
        params:
            name_column (str) : nombre de la culumna la cual bamos a buscar

        description: 
            nos permite buscar una columna a partir de su nombre o algun dato string que este en ella 
        """ 
        self.colum = {
            "value": 0,
            "list": list
        }
        #iteramos por cada fila en la hoja
        for row in list(self.sheet_names.rows):
            #iteramos por cada cellda en la fila
            for cell in row:

                #si el valor de la cellda es igual al dato que estamos buscando 
                if cell.value == name_column:

                    #asignamos los valores al diccionario
                    self.colum["list"] = row
                    self.colum["value"] = cell

        return self.colum


    #retorna un diccionario con un indice
    #EL INDICE es la posicon de la columna donde esta la TIIE
    def get_index_column_TIIE(self):
        self.data_colum = {
            "index": int,
        }

        row = self.get_row("TIIE")
        value_list = row["list"]
        for i in range(len(value_list)):
            if(row["value"] == value_list[i]):
                self.data_colum["index"] = i

        return self.data_colum

    # retorna dos listas
    # LA PRIMERA LIST es una lista de celdas las cuales tiene un valor asignado de TIIE
    # LA SEGUNDA LIST es una lista con las celdas las cuales no tiene un valor asignado

    def get_cell_TIIE(self):
        index_colum_TIIE = self.get_index_column_TIIE()
        values = []
        Cells_empty = []
        for cell in list(self.sheet_names.columns)[index_colum_TIIE["index"]]:
            if cell.value != None:
                values.append(cell)
            else:
                Cells_empty.append(cell)
        
        lists = {
        "Cells_whit_TIIE":values,
        "Cells_empty":Cells_empty
        } 

        return lists

        #retorna la siguiente celda en la clumna de las TIIE la cual este vacia
        
    def get_next_cell_empty(self):
        lists_cell_TIIE = self.get_cell_TIIE()
        list_cell_with_values_TIIE = lists_cell_TIIE["Cells_whit_TIIE"]
        last_cell_with_TIIE = list_cell_with_values_TIIE[-1]
        cell = self.get_cell_as_string(last_cell_with_TIIE)
        [letter ,cord_x, cord_y] = cell
        cord_y = int(cord_y)+1
        cell = letter + cord_x + str(cord_y)
        return cell

        # inserta un valor float en la siguiente celda de la column TIIE 
        
    def insert_TIIE_in_client(self,value: float, path_to_save: str = " "):
        cell = self.get_next_cell_empty()
        self.sheet_names[cell] = value
        self.Excel_document.save(self.path_file)
        print(self.sheet_names[cell].value)


if __name__ == "__main__":
    client = TIIE_file_edit_from_py(PATH)
    client.set_sheet_name("Alejandro Ochoa")
  
    # client.conver_string(1)
    cord = client.insert_TIIE_in_client(.9,"D:/desarrolloDeSofware/backend/python/TIIE/read_xlsx/Control de crédito mensual .xlsx")
    value = conver_value(12)
    tipe = value.conver_string()
    print(type(tipe))


